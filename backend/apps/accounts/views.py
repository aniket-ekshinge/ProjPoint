# accounts/views.py
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import authenticate, get_user_model
from rest_framework.permissions import IsAdminUser, IsAuthenticated, AllowAny
from rest_framework.authtoken.models import Token
from django.db import transaction
from django.contrib.auth.hashers import check_password

from .serializers import (UserSerializer, BulkUserUploadSerializer,
                          ChangePasswordSerializer, ProfileSerializer,AcademicYearUpdateSerializer)
import openpyxl

User = get_user_model()

class UserLoginView(APIView):
    """
    Endpoint for regular user login using email and password.
    Admin users are not allowed here.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        email = request.data.get('email')
        password = request.data.get('password')
        
        user = authenticate(request, email=email, password=password)
        if user is not None:
            if user.is_staff or user.is_superuser:
                return Response(
                    {'error': 'Admins must log in via the admin endpoint'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            token, created = Token.objects.get_or_create(user=user)
            
            return Response(
                {'message': 'User logged in successfully', 
                 'email': user.email,
                 'token': token.key},
                status=status.HTTP_200_OK
            )
        return Response({'error': 'Invalid credentials'},
                        status=status.HTTP_401_UNAUTHORIZED)

class AdminLoginView(APIView):
    """
    Admin login endpoint using username and password.
    This endpoint is left unchanged to work with your existing frontend.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        
        # Manually lookup user by username to preserve existing functionality.
        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            return Response({'error': 'Invalid credentials'},
                            status=status.HTTP_401_UNAUTHORIZED)

        if not check_password(password, user.password):
            return Response({'error': 'Invalid credentials'},
                            status=status.HTTP_401_UNAUTHORIZED)

        if user.is_staff or user.is_superuser:
            token, _ = Token.objects.get_or_create(user=user)
            return Response({
                'message': 'Admin logged in successfully',
                'username': user.username,
                'token': token.key
            }, status=status.HTTP_200_OK)
        return Response({'error': 'User does not have admin privileges'},
                        status=status.HTTP_403_FORBIDDEN)

class BulkUserUploadView(APIView):
    """
    Endpoint that allows an admin to upload an Excel file with student data.
    Only accessible by admin users.
    """
    permission_classes = [IsAdminUser]

    def post(self, request):
        serializer = BulkUserUploadSerializer(data=request.data)
        if serializer.is_valid():
            file_obj = serializer.validated_data['file']
            try:
                wb = openpyxl.load_workbook(file_obj)
                sheet = wb.active
            except Exception:
                return Response({'error': 'Invalid file format or error processing file'},
                                status=status.HTTP_400_BAD_REQUEST)
            
            users_created = []
            errors = []
            header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            # Expected header columns from the Excel file
            expected_columns = ['College_id', 'Full Name', 'Email', 'Phone No.', 'Academic Year', 'Branch', 'Password']
            if header != expected_columns:
                return Response({'error': f'Invalid header. Expected: {expected_columns}'},
                                status=status.HTTP_400_BAD_REQUEST)

            with transaction.atomic():
                for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        data = {
                            'college_id': row[0],
                            'full_name': row[1],
                            'email': row[2],
                            'phone_number': row[3],
                            'academic_year': row[4],
                            'branch': row[5],
                            'password': row[6],
                            # For regular login, we use email as username.
                            'username': row[2]
                        }
                        user_serializer = UserSerializer(data=data)
                        if user_serializer.is_valid():
                            user_serializer.save()
                            users_created.append(user_serializer.data)
                        else:
                            errors.append({f'Row {idx}': user_serializer.errors})
                    except Exception as e:
                        errors.append({f'Row {idx}': str(e)})

            response_data = {'created': users_created}
            if errors:
                response_data['errors'] = errors
            return Response(response_data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ChangePasswordView(APIView):
    """
    Endpoint for an authenticated user to change their password.
    Requires the user to supply the old password.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = ChangePasswordSerializer(data=request.data)
        user = request.user
        if serializer.is_valid():
            old_password = serializer.validated_data.get('old_password')
            new_password = serializer.validated_data.get('new_password')
            if not user.check_password(old_password):
                return Response({'error': 'Old password is incorrect'},
                                status=status.HTTP_400_BAD_REQUEST)
            user.set_password(new_password)
            user.save()
            return Response({'message': 'Password changed successfully'},
                            status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class UserAcademicYearUpdateView(APIView):
    """
    Allows a logged-in user to update their own academic year.
    """
    permission_classes = [IsAuthenticated]

    def put(self, request):
        user = request.user
        serializer = AcademicYearUpdateSerializer(user, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Academic year updated successfully'}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

class ProfileView(APIView):
    """
    Endpoint for users to view and update their profile.
    The profile data excludes the password.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = ProfileSerializer(request.user)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request):
        serializer = ProfileSerializer(request.user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Profile updated successfully', 'data': serializer.data},
                            status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
